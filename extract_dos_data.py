#!/usr/bin/env python3
"""
Extract full DOS (Days of Service) data for the data entry UI.
Builds rich per-employee records with all run/segment details.
"""

import json
import re
from pathlib import Path
from collections import defaultdict

import openpyxl


def is_valid_name(name: str) -> bool:
    if not name or not isinstance(name, str):
        return False
    name = name.strip()
    if not name or len(name) < 3:
        return False
    skip = {"primary driver name", "alternative driver name", "supervisors", "absent"}
    if name.lower() in skip:
        return False
    return "," in name and any(c.isalpha() for c in name)


def normalize_name(name: str) -> str:
    return " ".join(name.strip().split())


def _extract_from_sheet(ws, use_supervisor_sections: bool) -> tuple:
    """Extract from a worksheet. Returns (by_employee dict, date_str or None)."""
    COLS = {
        "paddle": 0,
        "block": 1,
        "planned_start": 2,
        "planned_end": 3,
        "planned_hrs": 4,
        "vehicle": 5,
        "actual_start": 6,
        "actual_end": 7,
        "actual_hrs": 8,
        "primary_name": 9,
        "primary_id": 10,
        "alt_name": 11,
        "alt_id": 12,
        "labels": 13,
        "driver_notes": 14,
        "internal_notes": 15,
        "cancelled": 16,
    }

    supervisor_header_row = None
    absent_header_row = None
    by_employee = defaultdict(lambda: {"runs": [], "in_supervisor_section": False, "in_operator_section": False})

    rows = list(ws.iter_rows(values_only=True))
    # Try to get date from sheet name (YYYY-MM-DD)
    date_str = None
    if ws.title and re.match(r"^\d{4}-\d{2}-\d{2}$", str(ws.title).strip()):
        date_str = str(ws.title).strip()

    for row_idx, row in enumerate(rows):
        row = list(row)
        paddle = str(row[COLS["paddle"]] or "").strip()

        if use_supervisor_sections:
            if paddle.upper() == "SUPERVISORS":
                supervisor_header_row = row_idx
                continue
            if paddle.upper() == "ABSENT":
                absent_header_row = row_idx
            after_supervisor = supervisor_header_row is not None and row_idx > supervisor_header_row
            before_absent = absent_header_row is None or row_idx < absent_header_row
            is_supervisor_row = after_supervisor and before_absent
        else:
            is_supervisor_row = False  # Raw format: no supervisor sections, all operators

        run = {
            "paddle": paddle,
            "block": str(row[COLS["block"]] or "").strip(),
            "planned_start": str(row[COLS["planned_start"]] or "").strip(),
            "planned_end": str(row[COLS["planned_end"]] or "").strip(),
            "planned_hrs": row[COLS["planned_hrs"]],
            "vehicle": str(row[COLS["vehicle"]] or "").strip(),
            "actual_start": str(row[COLS["actual_start"]] or "").strip(),
            "actual_end": str(row[COLS["actual_end"]] or "").strip(),
            "actual_hrs": row[COLS["actual_hrs"]],
            "labels": str(row[COLS["labels"]] or "").strip().replace("\n", " "),
            "driver_notes": str(row[COLS["driver_notes"]] or "").strip().replace("\n", " "),
            "internal_notes": str(row[COLS["internal_notes"]] or "").strip().replace("\n", " "),
            "cancelled": str(row[COLS["cancelled"]] or "").strip(),
        }
        # Convert numeric
        for k in ("planned_hrs", "actual_hrs"):
            try:
                run[k] = float(run[k]) if run[k] else None
            except (TypeError, ValueError):
                run[k] = None

        # Primary driver
        pname = str(row[COLS["primary_name"]] or "").strip()
        pid = str(row[COLS["primary_id"]] or "").strip()
        if is_valid_name(pname):
            key = (normalize_name(pname), pid)
            by_employee[key]["name"] = pname
            by_employee[key]["employee_id"] = pid
            by_employee[key]["runs"].append({**run, "role": "primary"})
            if is_supervisor_row:
                by_employee[key]["in_supervisor_section"] = True
            else:
                by_employee[key]["in_operator_section"] = True

        # Alt driver
        aname = str(row[COLS["alt_name"]] or "").strip()
        aid = str(row[COLS["alt_id"]] or "").strip()
        if is_valid_name(aname):
            key = (normalize_name(aname), aid)
            by_employee[key]["name"] = aname
            by_employee[key]["employee_id"] = aid
            by_employee[key]["runs"].append({**run, "role": "alt"})
            if is_supervisor_row:
                by_employee[key]["in_supervisor_section"] = True
            else:
                by_employee[key]["in_operator_section"] = True

    return by_employee, date_str


def extract_dos_data(excel_path: str, format: str = None) -> dict:
    """
    Extract full run data per employee for UI.
    format: 'standard' (Table 1 sheet, SUPERVISORS/ABSENT sections), 'raw' (first sheet, flat list), or None (auto-detect)
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    # Auto-detect format
    if format is None:
        format = "standard" if "Table 1" in wb.sheetnames else "raw"

    if format == "standard":
        ws = wb["Table 1"]
        use_supervisor_sections = True
    else:
        ws = wb.worksheets[0]
        use_supervisor_sections = False

    by_employee, date_str = _extract_from_sheet(ws, use_supervisor_sections)
    wb.close()

    # Build final list with employee_type
    employees = []
    for (name, eid), data in sorted(by_employee.items(), key=lambda x: (x[1]["name"].lower(), x[1]["employee_id"])):
        in_op = data["in_operator_section"]
        in_sup = data["in_supervisor_section"]
        if in_op and in_sup:
            emp_type = "both"
        elif in_sup:
            emp_type = "supervisor"
        else:
            emp_type = "operator"
        employees.append({
            "name": data["name"],
            "employee_id": data["employee_id"],
            "employee_type": emp_type,
            "skip": emp_type in ("supervisor", "both"),
            "runs": data["runs"],
        })
    return {"employees": employees, "date": date_str or "3/9/2026"}


def main():
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "/Users/david/Downloads/3.9.26.xlsx"
    fmt = sys.argv[2] if len(sys.argv) > 2 else None  # 'standard' or 'raw' to override auto-detect
    if not Path(path).exists():
        print(f"File not found: {path}")
        return 1
    data = extract_dos_data(path, format=fmt)
    out_dir = Path(__file__).resolve().parent
    out_json = out_dir / "dos_data.json"
    with open(out_json, "w") as f:
        json.dump(data, f, indent=2)
    print(f"Extracted {len(data['employees'])} employees to {out_json}")

    # Auto-build UI if build_ui exists
    build_ui = out_dir / "build_ui.py"
    if build_ui.exists():
        import subprocess
        subprocess.run([sys.executable, str(build_ui)], check=False)
    return 0


if __name__ == "__main__":
    exit(main() or 0)
